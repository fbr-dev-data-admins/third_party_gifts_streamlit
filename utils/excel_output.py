"""Excel file generation with conditional formatting."""

import io
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule


HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
PALE_YELLOW_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F2AC57", end_color="F2AC57", fill_type="solid")
PALE_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def create_import_excel(
    df: pd.DataFrame,
    benevity_rows: Optional[set] = None,
    benevity_reason_rows: Optional[set] = None,
    stale_cache_rows: Optional[set] = None
) -> bytes:
    """
    Create the unified import Excel file with conditional formatting.

    Args:
        df: DataFrame with unified import data
        benevity_rows: Set of row indices (0-based, data rows) that are from Benevity
        benevity_reason_rows: Set of row indices where Benevity Reason is non-standard

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Gift Import"

    benevity_rows = benevity_rows or set()
    benevity_reason_rows = benevity_reason_rows or set()
    stale_cache_rows = stale_cache_rows or set()

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)

            if r_idx == 0:
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")

    columns = list(df.columns)
    state_col_idx = columns.index("State") + 1 if "State" in columns else None
    zip_col_idx = columns.index("ZIP") + 1 if "ZIP" in columns else None
    amount_col_idx = columns.index("Gift Amount") + 1 if "Gift Amount" in columns else None
    ref_col_idx = columns.index("Gift Reference") + 1 if "Gift Reference" in columns else None
    first_name_col_idx = columns.index("First Name") + 1 if "First Name" in columns else None
    last_name_col_idx = columns.index("Last Name") + 1 if "Last Name" in columns else None
    constituent_id_col_idx = columns.index("Constituent ID") + 1 if "Constituent ID" in columns else None

    if zip_col_idx:
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=zip_col_idx)
            cell.number_format = "@"

    if state_col_idx and zip_col_idx:
        state_col = get_column_letter(state_col_idx)
        zip_col = get_column_letter(zip_col_idx)

        state_range = f"{state_col}2:{state_col}{len(df) + 1}"
        formula = f'AND(${state_col}2="",${zip_col}2<>"")'

        ws.conditional_formatting.add(
            state_range,
            FormulaRule(formula=[formula], fill=YELLOW_FILL)
        )

    for row_idx in benevity_rows:
        excel_row = row_idx + 2
        if amount_col_idx:
            try:
                amount = df.iloc[row_idx]["Gift Amount"]
                if pd.notna(amount) and float(amount) >= 1000:
                    ws.cell(row=excel_row, column=amount_col_idx).fill = ORANGE_FILL
            except (ValueError, TypeError):
                pass

    for row_idx in benevity_reason_rows:
        excel_row = row_idx + 2
        if ref_col_idx:
            ws.cell(row=excel_row, column=ref_col_idx).fill = YELLOW_FILL

    for row_idx in stale_cache_rows:
        excel_row = row_idx + 2
        for c_idx in range(1, len(columns) + 1):
            ws.cell(row=excel_row, column=c_idx).fill = PALE_GREEN_FILL
            
    for row_idx in range(len(df)):
        excel_row = row_idx + 2
        constituent_id = df.iloc[row_idx]["Constituent ID"] if "Constituent ID" in df.columns else ""
        is_non_anonymous = pd.notna(constituent_id) and str(constituent_id).strip() != ""

        if first_name_col_idx:
            first_name = df.iloc[row_idx]["First Name"] if "First Name" in df.columns else ""
            first_name_blank = pd.isna(first_name) or str(first_name).strip() == ""
            first_name_has_space = pd.notna(first_name) and " " in str(first_name)
            if first_name_has_space or (is_non_anonymous and first_name_blank):
                ws.cell(row=excel_row, column=first_name_col_idx).fill = PALE_YELLOW_FILL

        if last_name_col_idx:
            last_name = df.iloc[row_idx]["Last Name"] if "Last Name" in df.columns else ""
            last_name_blank = pd.isna(last_name) or str(last_name).strip() == ""
            last_name_has_space = pd.notna(last_name) and " " in str(last_name)
            if last_name_has_space or (is_non_anonymous and last_name_blank):
                ws.cell(row=excel_row, column=last_name_col_idx).fill = PALE_YELLOW_FILL

    for col_idx, column in enumerate(columns, 1):
        max_length = len(str(column))
        for row_idx in range(len(df)):
            try:
                cell_value = str(df.iloc[row_idx, col_idx - 1])
                max_length = max(max_length, len(cell_value))
            except (IndexError, TypeError):
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_grants_excel(grants_data: dict[str, tuple[list, pd.DataFrame]]) -> bytes:
    """
    Create the grants file with source headers and data stacked.

    Args:
        grants_data: Dict mapping source name to (original_headers, grants_df)

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Grants"

    current_row = 1

    for source_name, (headers, grants_df) in grants_data.items():
        if grants_df.empty:
            continue

        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
        current_row += 1

        for _, row in grants_df.iterrows():
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=current_row, column=c_idx, value=value)
            current_row += 1

        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
